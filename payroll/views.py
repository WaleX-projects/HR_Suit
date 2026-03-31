from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from employees.models import Employee
from attendance.models import Holiday, Attendance
from companies.models import Company 
from payroll.utils import get_working_days
import csv
from django.http import HttpResponse
from django.http import HttpResponse
from attendance.models import Holiday
from employees.models import Employee

from .utils import get_working_days, calculate_attendance
from .models import PayrollRun, Payslip
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse


from .models import (
    
    PositionSalary,
    SalaryComponent,
    PayrollRun,
    
    
    
)
from .serializers import (
    EmployeeSerializer,
    PositionSalarySerializer,
    SalaryComponentSerializer,
    PayrollRunSerializer,
    AttendanceSerializer,
    HolidaySerializer,
)

from .utils import generate_payroll


# =========================
# BASIC CRUD
# =========================

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer


class PositionSalaryViewSet(viewsets.ModelViewSet):
    queryset = PositionSalary.objects.all()
    serializer_class = PositionSalarySerializer


class SalaryComponentViewSet(viewsets.ModelViewSet):
    queryset = SalaryComponent.objects.all()
    serializer_class = SalaryComponentSerializer


class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer


class HolidayViewSet(viewsets.ModelViewSet):
    queryset = Holiday.objects.all()
    serializer_class = HolidaySerializer


class PayrollRunViewSet(viewsets.ModelViewSet):
    queryset = PayrollRun.objects.all()
    serializer_class = PayrollRunSerializer
    
    def get_queryset(self):
        company_id = self.request.user.company_id

        return PayrollRun.objects.filter(
            company_id=company_id
            ).prefetch_related(
                "payslips__employee",
                "payslips__items"
        )
  



    @action(detail=True, methods=["get"])
    def export_csv(self, request, pk=None):
        print("pk__analysis",pk)
        payroll = self.get_object()
    
        year = payroll.year
        month = payroll.month
        company = payroll.company
    
        employees = Employee.objects.filter(company=company)
    
        working_days = get_working_days(company, year, month)
        total_working_days = len(working_days)
    
        # =========================
        # CREATE WORKBOOK
        # =========================
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Report"
    
        # =========================
        # STYLES
        # =========================
        bold_big = Font(size=20, bold=True)
        bold = Font(size=12, bold=True)
    
        center = Alignment(horizontal="center", vertical="center")
    
        header_fill = PatternFill("solid", fgColor="DDDDDD")
    
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
        # =========================
        # HEADER SECTION
        # =========================
        ws.merge_cells("A1:I1")
        ws.merge_cells("A2:I2")
    
        ws["A1"] = company.name.upper()
        ws["A2"] = f"PAYROLL REPORT - {month}/{year}"
    
        ws["A1"].font = bold_big
        ws["A1"].alignment = center
    
        ws["A2"].font = bold
        ws["A2"].alignment = center
    
        ws.append([])
    
        # =========================
        # TABLE HEADER
        # =========================
        headers = [
            "Employee",
            "Position",
            "Working Days",
            "Absent Days",
            "present_Days",
            "Daily Salary",
            "Basic Salary",
            "Allowance",
            "Deduction",
            "Net Salary",
            "Account No",
            "Bank name",
            "Account name",
            "Account type"
        ]
    
        ws.append(headers)
    
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=col)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
    
        # =========================
        # DATA ROWS
        # =========================
        row_num = 5
    
        for employee in employees:
    
            payslip = payroll.payslips.filter(employee=employee).first()
            if not payslip:
                continue
    
            absent_days = calculate_attendance(employee, working_days)
            present_days = total_working_days - absent_days
    
            daily_salary = (
                payslip.basic_salary / total_working_days
                if total_working_days else 0
            )
    
            row = [
                f"{employee.first_name} {employee.last_name}",
                employee.position.title if employee.position else "",
                total_working_days,
                absent_days,
                present_days,
                round(daily_salary, 2),
                payslip.basic_salary,
                payslip.total_allowance,
                payslip.total_deduction,
                payslip.net_salary,
                employee.bank_account_number,
                employee.bank_name,
                employee.bank_account_name,
                employee.bank_account_type
            ]
            
    
            ws.append(row)
    
            # Add borders + alignment
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = center
    
            row_num += 1
    
        # =========================
        # RESPONSE FILE
        # =========================
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
        response["Content-Disposition"] = (
            f'attachment; filename="payroll_{company.name}_{month}_{year}.xlsx"'
        )
    
        wb.save(response)
    
        return response
        
    @action(detail=True, methods=["post"])
    def mark_paid(self, request, pk=None):
        payroll = self.get_object()

        try:
            
            payroll.mark_paid()
            return Response({"message": "Payroll approved"})
        except Exception as e:
            return Response({"error": str(e)}, status=400)
            
    
    @action(detail=True, methods=["post"])
    def process(self, request, pk=None):
        payroll = self.get_object()

        try:
            payroll.process()
            return Response({"message": "Payroll approved"})
        except Exception as e:
            return Response({"error": str(e)}, status=400)

    # =========================
    # RUN PAYROLL ENDPOINT
    # =========================
    @action(detail=False, methods=["post"])
    def run(self, request):
        company_id = request.user.company_id
        
        month = request.data.get("month")
        year = request.data.get("year")
        print(company_id, month, year)

        if not all([company_id, month, year]):
            return Response(
                {"error": "company, month, year required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        

        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return Response(
                {"error": "Company not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        payroll = generate_payroll(company, int(year), int(month))

        return Response(
            PayrollRunSerializer(payroll).data,
            status=status.HTTP_200_OK
        )
        
        
        
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.shortcuts import get_object_or_404

from employees.models import Employee
from .models import Payslip
from .serializers import PayslipSerializer


@api_view(["GET"])
def employee_payslips(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    payslips = Payslip.objects.filter(employee=employee).select_related("payroll").order_by("-created_at")
    print("payslip",payslips)
    serializer = PayslipSerializer(payslips, many=True)
    print("serialiser",serializer.data)
    return Response(serializer.data)     